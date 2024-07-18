# from .main import ALGORITHM, SECRET_KEY, ACCESS_TOKEN_EXPIRE_MINUTES
from datetime import datetime, timedelta, timezone, date
from sqlalchemy.ext.asyncio import AsyncSession
from jose import JWTError, jwt
from dotenv.main import load_dotenv
from fastapi import Request, Depends, HTTPException
import os
from .users import Users, Products, Region, UserProductPlan
from .doctors import DoctorCategory, Speciality, MedicalOrganization, Doctor, DoctorMonthlyPlan, DoctorFact
from .pharmacy import Reservation
from typing import Annotated
from .database import get_db
from fastapi.security import HTTPBearer
from openpyxl import Workbook, load_workbook
import shutil
from fastapi.responses import FileResponse
import os
import subprocess
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
import calendar


load_dotenv()


SECRET_KEY = os.environ['SECRET_KEY']
ALGORITHM = os.environ['ALGORITHM']
ACCESS_TOKEN_EXPIRE_MINUTES = 30

auth_header = HTTPBearer()


def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt


async def validate_token(request: Request):
    token = request.headers.get("Authorization")
    if not token:
        raise HTTPException(status_code=401, detail="Token is missing")
    try:
        token = token.split("Bearer ")[1]
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        id: str = payload.get("sub")
        if id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except (jwt.JWTError, IndexError):
        raise HTTPException(status_code=401, detail="Could not validate credentials")
    return id


async def get_current_user(id: Annotated[str, Depends(validate_token)], db: AsyncSession = Depends(get_db)):
    user = await db.get(Users, int(id))
    if user is None:
        raise HTTPException(
            status_code=401,
            detail="Could not validate credentials",
            headers={"WWW-Authenticate": "Bearer"},
        )
    return user


async def check_if_user_already_exists(username: str, db: AsyncSession):
    result = await db.execute(select(Users).where(Users.username==username))
    db_user = result.all()
    if db_user:
        raise HTTPException(
            status_code=400,
            detail="Username already exists"
        )
    return False


async def get_user(user_id: int, db: AsyncSession):
    user = await db.get(Users, user_id)
    if not user:
        raise HTTPException(status_code=400, detail="There isn't user with this id")
    return user 


async def write_excel(reservation_id: int, db: AsyncSession):
    source_excel_file = 'app/report/Book.xlsx'
    destination_excel_file = 'app/report/report.xlsx'
    sheet_name = 'Sheet1'
    # shutil.copy2(source_excel_file, destination_excel_file)
    import subprocess
    subprocess.call(['cp', 'app/report/Book.xlsx', 'app/report/report.xlsx'])
    destination_wb = load_workbook(destination_excel_file)
    destination_sheet = destination_wb[sheet_name]
    result = await db.execute(select(Reservation).options(selectinload(Reservation.pharmacy)).where(Reservation.id==reservation_id))
    reservation = result.scalar()
    destination_sheet['E2'] = reservation.discount
    destination_sheet['E6'] = reservation.pharmacy.company_name
    destination_sheet['E7'] = reservation.pharmacy.inter_branch_turnover
    count = 9
    for product in reservation.products:
        data_to_write = {
            f'D{count}' : product.product.name,
            f'E{count}' : product.quantity,
            f'F{count}' : product.reservation_price,
            f'G{count}' : product.reservation_discount_price,
            f'H{count}' : product.quantity * product.reservation_price
        }
        count += 1
        for cell_address, value in data_to_write.items():
            destination_sheet[cell_address] = value
    destination_sheet['H31'] = reservation.total_amount
    destination_sheet['G32'] = f"{reservation.discount}% скидка билан"
    destination_sheet['H32'] = reservation.total_payable
    destination_sheet['G33'] = "12% ндс билан"
    destination_sheet['H33'] = reservation.total_payable + 0.12 * reservation.total_payable
    filename = reservation.pharmacy.company_name + '_' + reservation.pharmacy.inter_branch_turnover + '_' + str(date.today()) + "_" + reservation.manufactured_company.name + '.xlsx'
    destination_wb.save(destination_excel_file)
    destination_wb.close()
    return FileResponse(filename=filename, path="app/report/report.xlsx")


PRODUCT_EXCEL_DICT = {
        1 : "N",        2 : "O",        6 : "N",        7 : "CA",        25 : "P",
        8 : "Q",        4 : "R",        9 : "S",        10 : "T",        5 : "U",
        11 : "V",        12 : "W",        13 : "X",        26 : "Y",        27 : "AT",
        14 : "AA",        15 : "AB",        18 : "AC",        28 : "AR",        19 : "AE",
        20 : "AF",        21 : "AG",        22 : "AH",        29 : "AI",        23 : "AJ",
        30 : "AP",        24 : "AL",        16 : "AM",        17 : "AN",        31:"AO",
        33: "AQ",        
}

PRODUCT_EXCEL_DICT2 = {
        1 : "AP",        2 : "AQ",        6 : "AP",        7 : "CA",        25 : "AR",
        8 : "AS",        4 : "AT",        9 : "AU",        10 : "AV",        5 : "AW",
        11 : "AX",        12 : "AY",        13 : "AZ",        26 : "BA",        27 : "CB",
        14 : "BC",        15 : "BD",        18 : "BE",        28 : "BZ",        19 : "BG",
        20 : "BH",        21 : "BI",        22 : "BJ",        29 : "BK",        23 : "BL",
        30 : "BX",        24 : "BN",        16 : "BO",        17 : "BP",        31: "BW",
        33: "BY"
}


async def write_proccess_to_excel(month: int, db: AsyncSession):
    year = datetime.now().year
    month = datetime.now().month if month is None else month
    num_days = calendar.monthrange(year, month)[1]
    start_date = date(year, month, 1)  
    end_date = date(year, month, num_days)
    source_excel_file = 'app/report/Base_Doc.xlsx'
    destination_excel_file = 'app/report/base_doctor.xlsx'
    sheet_name = 'База Врачей Актуальная'
    subprocess.call(['cp', source_excel_file, destination_excel_file])
    destination_wb = load_workbook(destination_excel_file)
    ws = destination_wb.active
    destination_sheet = destination_wb[sheet_name]
    count = 7
    data_to_write = {}
    result = await db.execute(select(Users).options(selectinload(Users.region)).filter(Users.status=="product_manager"))
    product_mamnagers = result.scalars().all()
    pm_names = {}
    for pm in product_mamnagers:
        result = await db.execute(select(Users).options(selectinload(Users.region)).filter(Users.status=="medical_representative", Users.product_manager_id==pm.id))
        med_reps = result.scalars().all()
        med_rep_quantity = len(med_reps)
        if med_rep_quantity > 1:
            ws.merge_cells(f"B{count}:B{count+med_rep_quantity-1}") 
        pm_names[f'B{count}'] = pm.full_name
        doctor_count = 0
        for user in med_reps:
            USER_PLAN_COMPLEATED = {
                    "AV" : 0,        "AW" : 0,
                    "AX" : 0,        "AY" : 0,        "AZ" : 0,        "BA" : 0,        "BB" : 0,
                    "BC" : 0,        "BD" : 0,        "BE" : 0,        "BF" : 0,        "BG" : 0,
                    "BH" : 0,        "BI" : 0,        "BJ" : 0,        "BK" : 0,        "BL" : 0,
                    "BM" : 0,        "BN" : 0,        "BO" : 0,        "BP" : 0,
                    "BQ": 0, "BR":0, "BS":0, "BT": 0, "BU":0, "BV":0, "BW":0, "BX":0, "BY":0, "BZ":0, "CA":0, "CB":0
            }
            data_to_write[f"C{count}"] = user.full_name
            data_to_write[f"H{count}"] = user.region.name
            result = await db.execute(select(UserProductPlan).filter(UserProductPlan.med_rep_id==user.id, UserProductPlan.plan_month>=start_date, UserProductPlan.plan_month<=end_date))
            user_plan = {}
            user_plan_sum = 0
            user_plan_compleated_sum = 0
            for prd in result.scalars().all():
                user_plan[f"{PRODUCT_EXCEL_DICT[prd.product_id]}{count}"] = prd.amount
                user_plan_sum += prd.amount * prd.price
            data_to_write.update(user_plan)
            med_rep_count = count 
            count += 1
            result = await db.execute(select(Doctor).options(selectinload(Doctor.pharmacy)).filter(Doctor.med_rep_id==user.id))
            for doctor in result.scalars().all():
                pharmacies = doctor.pharmacy
                pharmacy_quantity = len(pharmacies)
                result = await db.execute(select(DoctorMonthlyPlan).filter(DoctorMonthlyPlan.doctor_id == doctor.id, DoctorMonthlyPlan.date>=start_date, DoctorMonthlyPlan.date<=end_date))
                doctor_products = result.scalars().all()
                if pharmacy_quantity > 1:
                    for d_prd in doctor_products:
                        ws.merge_cells(f"{PRODUCT_EXCEL_DICT[d_prd.product_id]}{count}:{PRODUCT_EXCEL_DICT[d_prd.product_id]}{count+pharmacy_quantity-1}")
                    ws.merge_cells(f"AU{count}:AU{count+pharmacy_quantity-1}")
                    ws.merge_cells(f"CC{count}:CC{count+pharmacy_quantity-1}")
                    ws.merge_cells(f"CD{count}:CD{count+pharmacy_quantity-1}")
                ph_compleated_sum = 0
                row = count  
                for pharmacy in pharmacies:
                    doctor_fact = None
                    # result = await db.execute(select(DoctorMonthlyPlan).filter(DoctorMonthlyPlan.doctor_id == doctor.id, DoctorMonthlyPlan.date>=start_date, DoctorMonthlyPlan.date<=end_date))
                    doctor_plan = dict()
                    doctor_plan_sum = 0
                    for product in doctor_products:
                        doctor_count += 1
                        doctor_plan[f"{PRODUCT_EXCEL_DICT[product.product_id]}{row}"] = product.monthly_plan
                        doctor_plan_sum += product.monthly_plan * product.price
                        result = await db.execute(select(DoctorFact).filter(DoctorFact.doctor_id == doctor.id, DoctorFact.product_id == product.product_id, DoctorFact.pharmacy_id == pharmacy.id, DoctorFact.date>=start_date, DoctorFact.date<=end_date))
                        tmp = result.scalars().first()
                        if tmp is not None:
                            doctor_fact = tmp
                            USER_PLAN_COMPLEATED[f"{PRODUCT_EXCEL_DICT2[product.product_id]}"] += doctor_fact.fact
                            user_plan_compleated_sum += doctor_fact.fact * doctor_fact.price
                    data_to_write[f'D{count}'] = doctor.full_name
                    data_to_write[f'E{count}'] = doctor.contact1
                    data_to_write[f'F{count}'] = doctor.speciality.name
                    data_to_write[f'G{count}'] = doctor.medical_organization.name
                    data_to_write[f'H{count}'] = doctor.medical_organization.region.name
                    # data_to_write[f'I{count}'] = doctor_count
                    data_to_write[f'J{count}'] = doctor.category.name
                    data_to_write[f'K{count}'] = pharmacy.company_name
                    data_to_write[f'L{count}'] = pharmacy.contact1
                    data_to_write[f'AU{row}'] = doctor_plan_sum
                    data_to_write.update(doctor_plan)
                    if doctor_fact is not None:
                        data_to_write[f'{PRODUCT_EXCEL_DICT2[doctor_fact.product_id]}{count}'] = doctor_fact.fact
                        ph_compleated_sum += doctor_fact.fact * doctor_fact.price
                        # data_to_write[f'CC{row}'] = doctor_fact.fact * doctor_fact.price
                        # data_to_write[f"CD{row}"] = round(doctor_fact.fact * doctor_fact.price / doctor_plan_sum, 2)
                    count += 1
                data_to_write[f'CC{row}'] = ph_compleated_sum
                data_to_write[f"CD{row}"] = round(ph_compleated_sum / doctor_plan_sum, 2) if doctor_plan_sum != 0 else 0
            for key in USER_PLAN_COMPLEATED.keys():
                data_to_write[f"{key}{med_rep_count}"] = USER_PLAN_COMPLEATED[key]
            data_to_write[f"AU{med_rep_count}"] = user_plan_sum
            data_to_write[f"CC{med_rep_count}"] = user_plan_compleated_sum
            data_to_write[f"CD{med_rep_count}"] = round(user_plan_compleated_sum / user_plan_sum if user_plan_sum else 0, 2)
    data_to_write.update(pm_names)
    for cell_address, value in data_to_write.items():
        destination_sheet[cell_address] = value
    destination_wb.save(destination_excel_file)
    destination_wb.close()
    filename = 'BASE_DOCTOR.xlsx'
    return FileResponse(filename=filename, path=destination_excel_file)


async def get_doctor_or_404(id, db):
    obj = await db.get(Doctor, id)
    if (not obj) or (obj.deleted == True):
        raise HTTPException(status_code=404, detail=f"Doctor not found")
    return obj

